import pandas as pd
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils import get_column_letter
import os

class ExcelHandler:
    def __init__(self):
        pass

    def get_preview(self, file_path: str, num_rows: int = 5) -> dict:
        """
        엑셀 파일의 첫 N행을 미리보기로 반환합니다.
        
        Args:
            file_path: 엑셀 파일 경로
            num_rows: 미리보기할 행 수 (기본값: 5)
            
        Returns:
            Dict: {
                'columns': ['A', 'B', 'C', ...],
                'headers': ['상품코드', '상품명', ...],  # 1행 데이터
                'preview_rows': [['P001', '청소용 수세미', ...], ...]
            }
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")
        
        # 헤더 없이 읽기
        df = pd.read_excel(file_path, header=None, nrows=num_rows)
        
        # 열 문자 생성 (A, B, C, ..., Z, AA, AB, ...)
        columns = [get_column_letter(i + 1) for i in range(len(df.columns))]
        
        # 첫 행은 헤더로 사용
        headers = df.iloc[0].fillna("").astype(str).tolist() if len(df) > 0 else []
        
        # 나머지 행은 미리보기 데이터
        preview_rows = df.iloc[1:].fillna("").astype(str).values.tolist() if len(df) > 1 else []
        
        return {
            "columns": columns,
            "headers": headers,
            "preview_rows": preview_rows
        }

    def load_excel(self, file_path: str, product_name_col: str, keyword_col: str = None) -> list:
        """
        엑셀 파일을 읽어와서 처리할 데이터 리스트를 반환합니다.
        
        Args:
            file_path: 엑셀 파일 경로
            product_name_col: 상품명이 있는 열 문자 (예: 'A')
            keyword_col: (옵션) 기존 키워드가 있는 열 문자 (예: 'E')
            
        Returns:
            List[Dict]: [{'row_index': 2, 'product_name': '...', 'input_keyword': '...'}, ...]
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")

        # 열 문자를 인덱스로 변환 (0-based for pandas iloc)
        try:
            p_col_idx = column_index_from_string(product_name_col) - 1
            k_col_idx = (column_index_from_string(keyword_col) - 1) if keyword_col else None
        except ValueError:
            raise ValueError("올바르지 않은 열 문자입니다. (예: 'A', 'AB')")

        df = pd.read_excel(file_path, header=None) # 헤더 없이 읽어서 0행부터 처리 (또는 사용자가 지정?)
        # 보통 1행은 헤더이므로 2행(Index 1)부터 데이터라고 가정.
        # 하지만 pandas는 header=0(기본값)이면 1행을 헤더로 씀.
        # 여기서는 header=0으로 읽고 처리하는게 안전.
        
        df = pd.read_excel(file_path, header=0)
        
        data_list = []
        for idx, row in df.iterrows():
            # row index는 0부터 시작하지만 엑셀 행번호는 header(1행) + 1 + idx + 1 = idx + 2
            excel_row_num = idx + 2
            
            # iloc을 사용하여 열 인덱스로 접근
            p_name = str(row.iloc[p_col_idx]) if pd.notna(row.iloc[p_col_idx]) else ""
            k_word = ""
            if k_col_idx is not None:
                k_word = str(row.iloc[k_col_idx]) if pd.notna(row.iloc[k_col_idx]) else ""
                
            data_list.append({
                'row_index': excel_row_num,
                'product_name': p_name,
                'input_keyword': k_word
            })
            
        return data_list

    def save_results(self, file_path: str, results: list, column_mapping: dict):
        """
        결과를 원본 엑셀 파일에 사용자가 지정한 열에 저장합니다.
        (openpyxl을 사용하여 기존 서식 유지하며 쓰기)
        
        Args:
            file_path: 원본 엑셀 파일 경로
            results: 처리 결과 리스트
            column_mapping: 열 매핑 정보 (refined_product_name, keyword, category)
        """
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 열 문자를 인덱스로 변환
        refined_name_col_idx = column_index_from_string(column_mapping.get('refined_product_name', 'B'))
        keyword_col_idx = column_index_from_string(column_mapping.get('keyword', 'E'))
        category_col_idx = column_index_from_string(column_mapping.get('category', 'F'))
        
        # 데이터 쓰기 (헤더는 이미 있다고 가정, 또는 보존)
        for item in results:
            row_idx = item['row_index']
            ws.cell(row=row_idx, column=refined_name_col_idx, value=item.get('refined_name', ''))
            ws.cell(row=row_idx, column=keyword_col_idx, value=item.get('keywords', ''))
            ws.cell(row=row_idx, column=category_col_idx, value=item.get('category_code', ''))
            
        # 저장 (파일명 변경하여 원본 보존)
        new_path = file_path.replace(".xlsx", "_processed.xlsx")
        wb.save(new_path)
        print(f"결과 파일 저장 완료: {new_path}")
        return new_path
